from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Income Statement"

# Colors
HEADER_BG = "1F3864"    # dark navy
SECTION_BG = "D6E4F7"   # light blue
SUBROW_BG  = "EBF3FB"   # lighter blue
WHITE      = "FFFFFF"
DARK_GRAY  = "F2F2F2"
GREEN_FONT = "006400"
RED_FONT   = "8B0000"

# Fonts
header_font  = Font(bold=True, color=WHITE, size=10)
bold_font    = Font(bold=True, size=10)
normal_font  = Font(size=10)
italic_font  = Font(italic=True, size=9, color="595959")

# Borders
thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color="595959")
thin_border  = Border(bottom=Side(style="thin", color="BFBFBF"))
thick_border = Border(bottom=Side(style="medium", color="595959"))

def hdr_fill(color): return PatternFill("solid", fgColor=color)

# ── Column widths ──────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 48
for col in ["B","C","D","E","F","G"]:
    ws.column_dimensions[col].width = 14

# ── Headers ────────────────────────────────────────────────────────────────────
headers = ["Income Statement | TIKR.com", "31/12/21", "31/12/22", "31/12/23", "31/12/24", "31/12/25", "LTM"]
ws.append(headers)
for col_idx, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = hdr_fill(HEADER_BG)
    cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
ws.row_dimensions[1].height = 20

# ── Data rows ──────────────────────────────────────────────────────────────────
# Format: (label, [b21, c22, d23, e24, f25, g_ltm], style)
# style: "revenue" | "bold" | "pct" | "sub" | "normal" | "separator" | "section"
rows = [
    ("Revenues",                                  [None, None, 14240, 16747, 20074, 20074],       "bold"),
    ("Total Revenues",                            [None, None, 14240, 16747, 20074, 20074],       "revenue"),
    ("   % Change YoY",                           [None, None, None, 0.176, 0.199, None],         "pct_sub"),
    ("Cost of Goods Sold",                        [None, None, -4391, -5290, -6267, -6267],       "normal"),
    ("Gross Profit",                              [None, None, 9849, 11457, 13807, 13807],        "bold"),
    ("   % Change YoY",                           [None, None, None, 0.163, 0.205, None],         "pct_sub"),
    ("   % Gross Margins",                        [None, None, 0.692, 0.684, 0.688, 0.688],      "pct_sub"),
    ("Selling General & Admin Expenses",          [None, None, -5190, -5984, -6887, -6887],       "normal"),
    ("R&D Expenses",                              [None, None, -1414, -1615, -2052, -2052],       "normal"),
    ("Amortization of Goodwill and Intangibles",  [None, None, -828, -856, -897, -897],           "normal"),
    ("Total Operating Expenses",                  [None, None, -7432, -8455, -9836, -9836],       "bold"),
    ("Operating Income",                          [None, None, 2417, 3002, 3971, 3971],           "bold"),
    ("   % Change YoY",                           [None, None, None, 0.242, 0.323, None],         "pct_sub"),
    ("   % Operating Margins",                    [None, None, 0.170, 0.179, 0.198, 0.198],      "pct_sub"),
    ("Interest Expense",                          [None, None, -265, -305, -349, -349],           "normal"),
    ("Interest And Investment Income",            [None, None, 22, 107, 29, 29],                  "normal"),
    ("Income (Loss) On Equity Invest.",           [None, None, None, None, None, None],           "normal"),
    ("Currency Exchange Gains (Loss)",            [None, None, -41, -16, -12, -12],               "normal"),
    ("Other Non Operating Income (Expenses)",     [None, None, -15, -31, -34, -34],               "normal"),
    ("EBT Excl. Unusual Items",                   [None, None, 2118, 2757, 3605, 3605],           "bold"),
    ("Merger & Restructuring Charges",            [None, None, -69, -16, -101, -101],             "normal"),
    ("Impairment of Goodwill",                    [None, None, None, None, None, None],           "normal"),
    ("Gain (Loss) On Sale Of Investments",        [None, None, -59, -79, 139, 139],               "normal"),
    ("Gain (Loss) On Sale Of Assets",             [None, None, None, None, None, None],           "normal"),
    ("Asset Writedown",                           [None, None, -58, -386, -46, -46],              "normal"),
    ("In Process R&D Expenses",                   [None, None, None, None, None, None],           "normal"),
    ("Legal Settlements",                         [None, None, 111, None, -194, -194],            "normal"),
    ("Other Unusual Items",                       [None, None, -58, 5, -18, -18],                 "normal"),
    ("EBT Incl. Unusual Items",                   [None, None, 1985, 2281, 3385, 3385],           "bold"),
    ("Income Tax Expense",                        [None, None, -393, -436, -493, -493],           "normal"),
    ("Earnings From Continuing Operations",       [None, None, 1592, 1845, 2892, 2892],           "bold"),
    ("Extraordinary Item & Accounting Change",    [None, None, None, None, None, None],           "normal"),
    ("Net Income to Company",                     [None, None, 1592, 1845, 2892, 2892],           "bold"),
    ("Minority Interest",                         [None, None, 1, 8, 6, 6],                       "normal"),
    ("Net Income",                                [None, None, 1593, 1853, 2898, 2898],           "revenue"),
    ("Preferred Dividend and Other Adjustments",  [None, None, -23, None, None, None],            "normal"),
    ("Net Income to Common Incl Extra Items",     [None, None, 1570, 1853, 2898, 2898],           "bold"),
    ("   % Margins",                              [None, None, 0.110, 0.111, 0.144, 0.144],      "pct_sub"),
    ("Net Income to Common Excl. Extra Items",    [None, None, 1570, 1853, 2898, 2898],           "bold"),
    ("   % Margins",                              [None, None, 0.110, 0.111, 0.144, 0.144],      "pct_sub"),
    # Supplementary separator
    ("Supplementary Data:", [], "section"),
    ("Diluted EPS Excl Extra Items",              [None, None, 1.07, 1.25, 1.94, 1.94],          "bold"),
    ("   % Change YoY",                           [None, None, None, 0.168, 0.552, None],         "pct_sub"),
    ("Weighted Average Diluted Shares Outstanding",[None, None, 1463.50, 1485.90, 1494.50, 1494.50], "normal"),
    ("   % Change YoY",                           [None, None, None, 0.015, 0.006, None],         "pct_sub"),
    ("Weighted Average Basic Shares Outstanding", [None, None, 1453.00, 1471.50, 1480.40, 1480.40], "normal"),
    ("   % Change YoY",                           [None, None, None, 0.013, 0.006, None],         "pct_sub"),
    ("Payout Ratio %",                            [None, None, None, None, None, None],           "normal"),
    ("Basic EPS",                                 [None, None, 1.08, 1.26, 1.96, 1.96],          "normal"),
    ("EBITDA",                                    [None, None, 3613, 4271, 5339, 5339],           "bold"),
    ("   % Change YoY",                           [None, None, None, 0.182, 0.250, None],         "pct_sub"),
    ("EBITDAR",                                   [None, None, 3709, 4369, 5456, 5456],           "normal"),
    ("R&D Expense",                               [None, None, 1414, 1615, 2052, 2052],           "normal"),
    ("Effective Tax Rate %",                      [None, None, 0.198, 0.191, 0.146, 0.146],      "pct_sub"),
    # Price Factors separator
    ("Price Factors:", [], "section"),
    ("Market Cap",                                [60533.68, 66273.04, 84690.65, 131642.27, 141350.85, 141350.85], "bold"),
    ("Price Close (US$)",                         [42.48, 46.27, 57.81, 89.32, 95.35, 95.35],    "normal"),
    ("TEV",                                       [68048.68, 74869.04, 93280.65, 140646.27, 152356.85, 152356.85], "bold"),
]

NUM_FMT_INT  = '#,##0'
NUM_FMT_DEC  = '#,##0.00'
NUM_FMT_PCT  = '0.0%'
NUM_FMT_EPS  = '#,##0.00'

for row_data in rows:
    label, values, style = row_data
    r = ws.max_row + 1

    if style == "section":
        ws.append([label])
        cell = ws.cell(row=r, column=1)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = hdr_fill("2F5597")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 16
        # merge across all cols
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        continue

    row_vals = [label] + (values if values else [None]*6)
    ws.append(row_vals)

    # Label cell
    label_cell = ws.cell(row=r, column=1)

    # Style the label
    if style in ("revenue",):
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = hdr_fill("D6E4F7")
        label_cell.border = thick_border
    elif style == "bold":
        label_cell.font = bold_font
        label_cell.fill = hdr_fill(DARK_GRAY)
    elif style == "pct_sub":
        label_cell.font = italic_font
        label_cell.fill = hdr_fill(WHITE)
    else:
        label_cell.font = normal_font
        label_cell.fill = hdr_fill(WHITE)

    # Number cells
    for col_idx, val in enumerate(values, 2):
        cell = ws.cell(row=r, column=col_idx)
        if val is None:
            cell.value = None
        else:
            cell.value = val

        # formatting
        if style == "pct_sub":
            cell.number_format = NUM_FMT_PCT
            cell.font = italic_font
        elif label in ("Diluted EPS Excl Extra Items", "Basic EPS"):
            cell.number_format = NUM_FMT_EPS
            if style == "bold":
                cell.font = bold_font
            else:
                cell.font = normal_font
        elif label in ("Weighted Average Diluted Shares Outstanding",
                       "Weighted Average Basic Shares Outstanding"):
            cell.number_format = '#,##0.00'
            cell.font = normal_font
        elif label == "Price Close (US$)":
            cell.number_format = '#,##0.00'
            cell.font = normal_font
        else:
            cell.number_format = NUM_FMT_INT
            if style in ("revenue", "bold"):
                cell.font = bold_font
                if style == "revenue":
                    cell.fill = hdr_fill("D6E4F7")
                    cell.border = thick_border
                else:
                    cell.fill = hdr_fill(DARK_GRAY)
            else:
                cell.font = normal_font
                cell.fill = hdr_fill(WHITE)

        cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[r].height = 15

# ── Freeze panes ───────────────────────────────────────────────────────────────
ws.freeze_panes = "B2"

output_path = r"C:\Users\Ramongus\Desktop\Ramon\Trabajo\ProyectosPropios\TradingApp\IncomeStatement.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
